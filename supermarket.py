import os
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup


class ExtractData:
    def __init__(self, purchase_data_dir=None):
        self.purchase_data_dir = purchase_data_dir or os.getcwd()
        self.purchase_files_list = []
        self.products_set = set()
        self.products_url = None


    def get_file_names_from_dir(self):
        self.purchase_files_list = [f for f in os.listdir(self.purchase_data_dir) if
                                    os.path.isfile(os.path.join(self.purchase_data_dir, f))]


    def get_products_set(self):
        for file in self.purchase_files_list:
            # Открываем эксель файл
            workbook = openpyxl.load_workbook(os.path.join(self.purchase_data_dir, file))
            # Выбираем активный лист
            sheet = workbook.active
            for i in range(1, sheet.max_row + 1):
                value = sheet.cell(row=i, column=1).value
                if value and type(value) == str and len(value) >= 15:
                    self.products_set.add(value)


    def get_dataframe_product_url(self):
        data = []  # Инициализируем список для хранения данных
        print(len(self.products_set))
        # Проходимся по каждому продукту в множестве
        for product in self.products_set:
            print(product)
            # Отправляем запрос на поиск в Google
            response = requests.get("https://www.google.com/search", params={"q": "ашан " + product})

            # Используем BeautifulSoup для парсинга HTML и извлечения ссылки на первый результат поиска
            soup = BeautifulSoup(response.text, "html.parser")
            all_links = soup.find_all('a')
            auchan_link = None

            # Ищем первую ссылку, содержащую "auchan" в тексте
            link = "No results found"
            for link in all_links:
                if link.text.lower().find("auchan") != -1:
                    auchan_link = link.get('href')
                    break

            # Добавляем данные в список
            data.append([product, link])

            # Создаем датафрейм из списка данных
            self.products_url = pd.DataFrame(data, columns=['product', 'producturl'])
            return self.products_url


# Пример использования
# extractor = ExtractData()
# extractor.get_file_names_from_dir()
# print(extractor.purchase_files_list)